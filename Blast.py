from Bio.Seq import Seq
from Bio import SeqIO
from Bio.SeqRecord import SeqRecord
from Bio.Blast.Applications import NcbiblastnCommandline
from Bio.Blast.Applications import NcbimakeblastdbCommandline
from openpyxl import  load_workbook
import os

def find_and_set_seqr(part_id):
    wb = load_workbook('E:\\4-Undergraduate\\a-iGEM\\5-Software\\Final\\whole_collection_processed.xlsx')
    ws = wb['Sheet1']
    for cell in ws['A']:
        if cell.value == part_id:
            row = cell.row
            part_name = ws['B'+str(row)].value
            part_url = ws['D'+str(row)].value
            part_year = ws['H'+str(row)].value
            part_seq = ws['I'+str(row)].value
            if 'Part' in part_seq:
                part_seq = 'ATTATT'
            part_contents = ws['J'+str(row)].value
            part_status = ws['K'+str(row)].value
            part_stars = ws['M'+str(row)].value
            part_twins = ws['N'+str(row)].value
            part_assemble = ws['O'+str(row)].value
            part_usingparts = ws ['Q'+str(row)].value
            part_usedparts = ws['P'+str(row)].value
            part_len = ws['R'+str(row)].value
            part_type = ws['F'+str(row)].value
            part_team = ws['G'+str(row)].value
            part_release = ws['K'+ str(row)].value

            #注意如果sequence格子中的不是序列的报错行为

            target_seqr = SeqRecord(part_seq)
            target_seqr.id = part_id
            target_seqr.name = part_name
            target_seqr.description = part_contents
            target_seqr.annotations['url'] = part_url
            target_seqr.annotations['year'] = part_year
            target_seqr.annotations['status'] = part_status
            target_seqr.annotations['stars'] = part_stars
            target_seqr.annotations['twins'] = part_twins
            target_seqr.annotations['assemble'] = part_assemble
            target_seqr.annotations['using_parts'] = part_usingparts
            target_seqr.annotations['used_parts'] = part_usedparts
            target_seqr.annotations['len'] = part_len
            target_seqr.annotations['type'] = part_type
            target_seqr.annotations['team'] = part_team
            target_seqr.annotations['release'] = part_release
            return target_seqr

#目前只对05年有效，记得正式使用时做更改

def transform_fasta_database():
    wb = load_workbook('whole_collection_processed.xlsx')
    ws = wb['Sheet1']
    seqrlst = []
    replacer = []
    for cell in ws['A']:
        if cell.value != '' and cell.value != 'part_num':
                row = cell.row
                part_id = ws['A'+str(row)].value
                part_name = ws['B' + str(row)].value
                part_url = ws['D' + str(row)].value
                part_year = ws['H' + str(row)].value
                part_seq = ws['I' + str(row)].value

                if part_seq == 'NA':
                    part_seq = "attatt"
                    print(str(part_id) +'处理A')
                    print('now seq ='+ part_seq)
                if ('Part' in part_seq):
                    part_seq = "attatt"
                    print(str(part_id) +'处理B')
                    print('now seq ='+ part_seq)


                part_contents = ws['J' + str(row)].value
                try:
                    part_contents.encode(encoding='gbk')
                except:
                    replacer.append(part_contents)
                    part_contents = 'somethingwaitingtobereplaced0010'
                part_status = ws['K' + str(row)].value
                part_stars = ws['M' + str(row)].value
                part_twins = ws['N' + str(row)].value
                part_assemble = ws['O' + str(row)].value
                part_usingparts = ws['Q' + str(row)].value
                part_usedparts = ws['P' + str(row)].value
                part_len = ws['R' + str(row)].value
                part_type = ws['F' + str(row)].value
                part_team = ws['G' + str(row)].value
                part_release = ws['K' + str(row)].value

                # 注意如果sequence格子中的不是序列的报错行为
                #Imp
                target_seqr = SeqRecord(Seq(part_seq))
                target_seqr.id = part_id
                target_seqr.name = part_name
                #target_seqr.description = part_contents
                target_seqr.annotations['url'] = part_url
                target_seqr.annotations['year'] = part_year
                target_seqr.annotations['status'] = part_status
                target_seqr.annotations['stars'] = part_stars
                target_seqr.annotations['twins'] = part_twins
                target_seqr.annotations['assemble'] = part_assemble
                target_seqr.annotations['using_parts'] = part_usingparts
                target_seqr.annotations['used_parts'] = part_usedparts
                target_seqr.annotations['len'] = part_len
                target_seqr.annotations['type'] = part_type
                target_seqr.annotations['team'] = part_team
                target_seqr.annotations['release'] = part_release
                seqrlst.append(target_seqr)

    #print(seqrlst[0].seq)
    #print(len(seqrlst))
    #SeqIO.write(seqrlst[2], "D:\completed_collection\\test.faa", 'fasta')
    SeqIO.write(seqrlst, "whole_collection_fasta.fasta", "fasta")
    return 1


#blast需要安装ncbi本地blast
def do_blast(working_seqr):
    #ncbiblast安装的路径很关键，决定cmd参数，各种路径都不要有空格
    cline = NcbimakeblastdbCommandline(cmd=r"E:\\a-MyCode\\Extension\\blast-2.12.0+\\bin\\makeblastdb" , dbtype="nucl",out="temp_db",title="test",input_file="whole_collection_fasta.fasta")
    #print(cline)
    cline()
    out_path = str(working_seqr.id) + "_blastresult.xml"
    blastn_cline = NcbiblastnCommandline(cmd=r"E:\\a-MyCode\\Extension\\blast-2.12.0+\\bin\\blastn",query="whole_collection_fasta.fasta", db="temp_db", evalue=0.0000000000000000000001, outfmt = 5, out = out_path)
    #print(blastn_cline)
    blastn_cline()
    result_handle = open(out_path, encoding='gb18030', errors= 'ignore')
    blast_records=NCBIXML.parse(result_handle)
    E_VALUE_THRESH = 0.00000000000000000000000000000001  # 给出所有大于等于某一特定阈值的BLAST命中结果的一些汇总信息，建议取零
    for blast_record in blast_records:
        for alignment in blast_record.alignments:
            for hsp in alignment.hsps:
                if hsp.expect <= E_VALUE_THRESH:
                    result_path = 'BlastHistoryResult\\' +str(working_seqr.id) + '_BlastAnalysis'
                    with open(result_path, "a+") as blastanalysis:
                        blastanalysis.write('****Alignment****\n')
                        blastanalysis.write('sequence:'+alignment.title+"\n")
                        blastanalysis.write('ilength:'+ str(alignment.length)+"\n")
                        blastanalysis.write('e value:'+ str(hsp.expect)+"\n")
                        blastanalysis.write(hsp.query[0:75] + '...'+"\n")
                        blastanalysis.write(hsp.match[0:75] + '...'+"\n")
                        blastanalysis.write(hsp.sbjct[0:75] + '...'+"\n")
    os.remove('temp_db.ndb')
    os.remove('temp_db.nhr')
    os.remove('temp_db.nin')
    os.remove('temp_db.not')
    os.remove('temp_db.nsq')
    os.remove('temp_db.ntf')
    os.remove('temp_db.nto')
    result_handle.close()
    os.remove(out_path)

    #print('end')


if __name__ == '__main__':
    #set_database()
    # get_cite_parts()  #注意改completed_collection!加下划线连接
    # working_seqr = find_and_set_seqr('BBa_J07009')
    working_seqr = find_and_set_seqr('BBa_I15008')
    transform_fasta_database()  # 改了一下，现在不用改包也不会有encoding error
    do_blast(working_seqr)  #跑这个要谨慎，一跑一个多G就无了
    #set_database()
    #fp_get()  #跑这个前必须先跑过get_cite_parts获得数据库
    #get_relevant_parts('BBa_J13100')  #跑这个前必须先跑过fp_get获得数据库
    #return 0